"""
Excel Test Case Exporter
Generates professional Excel test case documents with 4 sheets:
  1. Summary        – Ticket info, test case statistics, generation metadata
  2. Test Cases     – Complete catalog (Test ID, Priority, Category, Title, Steps, Expected Results)
  3. QA Roadmap     – Test scenarios by category, coverage overview
  4. Coverage Analysis – Requirements coverage, identified gaps, clarification questions
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from datetime import datetime
from io import BytesIO
import json


class ExcelExporter:
    """
    Exports test cases to a professional Excel format
    """

    # Shared style constants
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    SECTION_FONT = Font(size=12, bold=True, color="1F4E79")
    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    PRIORITY_COLORS = {
        "P0": "FF4444",  # Red
        "P1": "FF8C00",  # Orange
        "P2": "FFD700",  # Yellow
        "P3": "90EE90",  # Light Green
        "P4": "87CEEB",  # Light Blue
    }

    def __init__(self):
        pass

    # ── public entry point ──────────────────────────────────────
    def export_test_cases_to_bytes(self, state: dict) -> BytesIO:
        wb = Workbook()

        self._create_summary_sheet(wb, state)
        self._create_test_cases_sheet(wb, state)
        self._create_qa_roadmap_sheet(wb, state)
        self._create_coverage_sheet(wb, state)

        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── helpers ─────────────────────────────────────────────────
    def _header_row(self, ws, headers: list, row: int = 1):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.THIN_BORDER

    def _section_heading(self, ws, row: int, text: str, merge_end_col: int = 2):
        cell = ws.cell(row, 1, text)
        cell.font = self.SECTION_FONT
        cell.fill = self.SECTION_FILL
        cell.border = self.THIN_BORDER
        if merge_end_col > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_end_col)
            for c in range(2, merge_end_col + 1):
                ws.cell(row, c).fill = self.SECTION_FILL
                ws.cell(row, c).border = self.THIN_BORDER
        return row + 1

    def _bullet_list(self, ws, row: int, items: list, icon: str = "•", fill=None) -> int:
        for item in items:
            ws.cell(row, 1, icon).alignment = Alignment(horizontal="center")
            c = ws.cell(row, 2, str(item))
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                c.fill = fill
            row += 1
        return row

    @staticmethod
    def _safe_str(value) -> str:
        if value is None:
            return ""
        if isinstance(value, (list, dict)):
            return json.dumps(value, indent=2, ensure_ascii=False)
        return str(value)

    # ═══════════════════════════════════════════════════════════
    #  SHEET 1 – SUMMARY
    # ═══════════════════════════════════════════════════════════
    def _create_summary_sheet(self, wb: Workbook, state: dict):
        ws = wb.active
        ws.title = "Summary"

        # Title banner
        ws.merge_cells("A1:D1")
        banner = ws["A1"]
        banner.value = "Ticket-to-Test AI — QA Execution Summary"
        banner.font = Font(size=16, bold=True, color="FFFFFF")
        banner.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        banner.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32
        for c in range(2, 5):
            ws.cell(1, c).fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        ticket = state.get("ticket_info", {})
        row = 3

        # — Ticket information ——————————————————————————————————
        row = self._section_heading(ws, row, "Ticket Information", 4)

        info_fields = [
            ("Ticket ID",            ticket.get("ticket_id", "N/A")),
            ("Title",               ticket.get("title", "N/A")),
            ("Type",                ticket.get("ticket_type", "N/A")),
            ("Priority",            ticket.get("priority", "N/A")),
            ("Status",              ticket.get("status", "N/A")),
        ]
        for label, value in info_fields:
            ws.cell(row, 1, label).font = Font(bold=True)
            ws.cell(row, 2, self._safe_str(value))
            row += 1

        # Description / Acceptance Criteria (may be long)
        desc = ticket.get("description", "")
        ac = ticket.get("acceptance_criteria", "")
        if desc:
            ws.cell(row, 1, "Description").font = Font(bold=True)
            ws.cell(row, 2, self._safe_str(desc)).alignment = Alignment(wrap_text=True)
            row += 1
        if ac:
            ws.cell(row, 1, "Acceptance Criteria").font = Font(bold=True)
            ws.cell(row, 2, self._safe_str(ac)).alignment = Alignment(wrap_text=True)
            row += 1

        row += 1

        # — Test case statistics ————————————————————————————————
        row = self._section_heading(ws, row, "Test Case Statistics", 4)

        test_cases = state.get("test_cases", [])
        priority_counts: Dict[str, int] = {}
        category_counts: Dict[str, int] = {}
        for tc in test_cases:
            p = tc.get("priority", "P2")
            cat = tc.get("category", "Other")
            priority_counts[p] = priority_counts.get(p, 0) + 1
            category_counts[cat] = category_counts.get(cat, 0) + 1

        ws.cell(row, 1, "Total Test Cases").font = Font(bold=True)
        ws.cell(row, 2, len(test_cases)).font = Font(bold=True, size=12)
        row += 1

        ws.cell(row, 1, "By Priority").font = Font(bold=True, italic=True)
        row += 1
        for p in sorted(priority_counts.keys()):
            ws.cell(row, 2, p)
            count_cell = ws.cell(row, 3, priority_counts[p])
            color = self.PRIORITY_COLORS.get(p, "FFFFFF")
            count_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            count_cell.alignment = Alignment(horizontal="center")
            row += 1

        row += 1
        ws.cell(row, 1, "By Category").font = Font(bold=True, italic=True)
        row += 1
        for cat in sorted(category_counts.keys()):
            ws.cell(row, 2, cat)
            ws.cell(row, 3, category_counts[cat]).alignment = Alignment(horizontal="center")
            row += 1

        row += 1

        # — Generation metadata ————————————————————————————————
        row = self._section_heading(ws, row, "Generation Metadata", 4)

        processing_time = state.get("processing_time", 0)
        try:
            processing_time = float(processing_time)
        except (TypeError, ValueError):
            processing_time = 0

        meta_fields = [
            ("Generated On",           datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Processing Time",        f"{processing_time:.1f} seconds" if processing_time else "N/A"),
            ("Requirements Extracted", len(state.get("extracted_requirements", []))),
            ("Coverage Gaps Found",    len(state.get("coverage_gaps", []))),
            ("Risk Areas Identified",  len(state.get("risk_areas", []))),
            ("Clarification Questions", len(state.get("clarification_questions", []))),
            ("Impacted Modules",       len(state.get("impacted_modules", []))),
            ("Dependencies",           len(state.get("dependencies", []))),
        ]
        for label, value in meta_fields:
            ws.cell(row, 1, label).font = Font(bold=True)
            ws.cell(row, 2, value)
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14

    # ═══════════════════════════════════════════════════════════
    #  SHEET 2 – TEST CASES
    # ═══════════════════════════════════════════════════════════
    def _create_test_cases_sheet(self, wb: Workbook, state: dict):
        ws = wb.create_sheet("Test Cases")

        headers = [
            "Test ID", "Priority", "Category", "Title",
            "Preconditions", "Test Steps", "Expected Result", "Test Data"
        ]
        self._header_row(ws, headers)

        test_cases = state.get("test_cases", [])
        for idx, tc in enumerate(test_cases, 2):
            # Test ID: use test_id if exists, else generate TC-{n}
            test_id = tc.get("test_id") or tc.get("id") or f"TC-{idx - 1}"
            ws.cell(idx, 1, self._safe_str(test_id)).alignment = Alignment(horizontal="center", vertical="top")

            # Priority with colour
            priority = tc.get("priority", "P2")
            p_cell = ws.cell(idx, 2, priority)
            color = self.PRIORITY_COLORS.get(priority, "FFFFFF")
            p_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            p_cell.alignment = Alignment(horizontal="center", vertical="top")

            ws.cell(idx, 3, tc.get("category", "")).alignment = Alignment(vertical="top")
            ws.cell(idx, 4, tc.get("title", "")).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(idx, 5, self._safe_str(tc.get("preconditions", ""))).alignment = Alignment(wrap_text=True, vertical="top")

            # Steps as numbered list
            steps = tc.get("test_steps", [])
            if isinstance(steps, list):
                steps_text = "\n".join(f"{i+1}. {s}" for i, s in enumerate(steps))
            else:
                steps_text = self._safe_str(steps)
            ws.cell(idx, 6, steps_text).alignment = Alignment(wrap_text=True, vertical="top")

            ws.cell(idx, 7, self._safe_str(tc.get("expected_result", ""))).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(idx, 8, self._safe_str(tc.get("test_data", ""))).alignment = Alignment(wrap_text=True, vertical="top")

        # Column widths
        widths = [14, 10, 18, 42, 30, 55, 35, 25]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

    # ═══════════════════════════════════════════════════════════
    #  SHEET 3 – QA ROADMAP
    # ═══════════════════════════════════════════════════════════
    def _create_qa_roadmap_sheet(self, wb: Workbook, state: dict):
        ws = wb.create_sheet("QA Roadmap")

        # Title banner
        ws.merge_cells("A1:C1")
        banner = ws["A1"]
        banner.value = "QA Execution Roadmap — Test Scenarios by Category"
        banner.font = Font(size=14, bold=True, color="FFFFFF")
        banner.fill = self.HEADER_FILL
        banner.alignment = Alignment(horizontal="center")
        for c in range(2, 4):
            ws.cell(1, c).fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        row = 3
        qa_roadmap = state.get("qa_roadmap", {})

        if qa_roadmap:
            for category, scenarios in qa_roadmap.items():
                # Category header
                row = self._section_heading(ws, row, str(category), 3)

                if isinstance(scenarios, list):
                    for scenario in scenarios:
                        ws.cell(row, 1, "•").alignment = Alignment(horizontal="center")
                        ws.cell(row, 2, str(scenario)).alignment = Alignment(wrap_text=True)
                        row += 1
                else:
                    ws.cell(row, 2, self._safe_str(scenarios)).alignment = Alignment(wrap_text=True)
                    row += 1

                row += 1  # blank row between categories
        else:
            ws.cell(row, 1, "No QA roadmap data available.")
            row += 1

        # Coverage overview section
        row += 1
        row = self._section_heading(ws, row, "Coverage Overview", 3)
        test_cases = state.get("test_cases", [])
        category_counts: Dict[str, int] = {}
        for tc in test_cases:
            cat = tc.get("category", "Other")
            category_counts[cat] = category_counts.get(cat, 0) + 1

        ws.cell(row, 1, "Category").font = Font(bold=True)
        ws.cell(row, 2, "Test Cases").font = Font(bold=True)
        row += 1
        for cat in sorted(category_counts.keys()):
            ws.cell(row, 1, cat)
            ws.cell(row, 2, category_counts[cat]).alignment = Alignment(horizontal="center")
            row += 1

        ws.cell(row, 1, "Total").font = Font(bold=True)
        ws.cell(row, 2, len(test_cases)).font = Font(bold=True)

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 80
        ws.column_dimensions["C"].width = 20

    # ═══════════════════════════════════════════════════════════
    #  SHEET 4 – COVERAGE ANALYSIS
    # ═══════════════════════════════════════════════════════════
    def _create_coverage_sheet(self, wb: Workbook, state: dict):
        ws = wb.create_sheet("Coverage Analysis")

        # Title banner
        ws.merge_cells("A1:B1")
        banner = ws["A1"]
        banner.value = "Coverage Analysis — Requirements, Gaps & Questions"
        banner.font = Font(size=14, bold=True, color="FFFFFF")
        banner.fill = self.HEADER_FILL
        banner.alignment = Alignment(horizontal="center")
        ws.cell(1, 2).fill = self.HEADER_FILL

        row = 3

        # — 4a. Requirements Coverage ———————————————————————————
        reqs = state.get("extracted_requirements", [])
        row = self._section_heading(ws, row, f"Requirements Coverage ({len(reqs)})", 2)
        if reqs:
            row = self._bullet_list(ws, row, reqs, "✓",
                                    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"))
        else:
            ws.cell(row, 2, "No explicit requirements extracted.")
            row += 1
        row += 1

        # — 4b. Acceptance Criteria Gaps ————————————————————————
        ac_gaps = state.get("acceptance_criteria_gaps", [])
        row = self._section_heading(ws, row, f"Acceptance Criteria Gaps ({len(ac_gaps)})", 2)
        if ac_gaps:
            row = self._bullet_list(ws, row, ac_gaps, "△",
                                    PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"))
        else:
            ws.cell(row, 2, "No acceptance criteria gaps detected.").fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            row += 1
        row += 1

        # — 4c. Identified Coverage Gaps ————————————————————————
        gaps = state.get("coverage_gaps", [])
        row = self._section_heading(ws, row, f"Identified Coverage Gaps ({len(gaps)})", 2)
        if gaps:
            row = self._bullet_list(ws, row, gaps, "⚠",
                                    PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"))
        else:
            ws.cell(row, 2, "No coverage gaps identified — Excellent coverage!").fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            row += 1
        row += 1

        # — 4d. Risk Areas ——————————————————————————————————————
        risks = state.get("risk_areas", [])
        row = self._section_heading(ws, row, f"Risk Areas ({len(risks)})", 2)
        if risks:
            row = self._bullet_list(ws, row, risks, "⬤",
                                    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        else:
            ws.cell(row, 2, "No risk areas identified.")
            row += 1
        row += 1

        # — 4e. Impacted Modules ————————————————————————————————
        modules = state.get("impacted_modules", [])
        if modules:
            row = self._section_heading(ws, row, f"Impacted Modules ({len(modules)})", 2)
            row = self._bullet_list(ws, row, modules, "◆")
            row += 1

        # — 4f. Dependencies ————————————————————————————————————
        deps = state.get("dependencies", [])
        if deps:
            row = self._section_heading(ws, row, f"Dependencies ({len(deps)})", 2)
            row = self._bullet_list(ws, row, deps, "→")
            row += 1

        # — 4g. Clarification Questions —————————————————————————
        questions = state.get("clarification_questions", [])
        row = self._section_heading(ws, row, f"Clarification Questions ({len(questions)})", 2)
        if questions:
            row = self._bullet_list(ws, row, questions, "?")
        else:
            ws.cell(row, 2, "No clarification needed.")
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 100


# ═══════════════════════════════════════════════════════════════
#  Public helpers
# ═══════════════════════════════════════════════════════════════
def export_to_excel_bytes(state: dict) -> BytesIO:
    """
    Export test cases to Excel in-memory.
    Returns BytesIO containing the .xlsx file.
    """
    return ExcelExporter().export_test_cases_to_bytes(state)


def get_excel_filename(state: dict) -> str:
    """
    Generate filename: TestCases_{TicketID}_{Timestamp}.xlsx
    """
    ticket_id = state.get("ticket_info", {}).get("ticket_id", "unknown")
    # Sanitise ticket ID for safe filenames
    safe_id = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in str(ticket_id))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"TestCases_{safe_id}_{timestamp}.xlsx"
